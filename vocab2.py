import tkinter as tk
import PyPDF2
from PIL import Image,ImageTk
from openpyxl import Workbook
import openpyxl
from openpyxl.utils import get_column_letter

root = tk.Tk()
canvas = tk.Canvas(root,width = 600 , height= 300)
canvas.grid(columnspan=3,rowspan= 6)

# instruction = tk.Label(root,text="Enrich your vocabulary!!",font="Raleway")
# instruction.grid(columnspan=3,column=0,row=1)
def check_word(a):
wb = openpyxl.load_workbook('vocab.xlsx')
# wb = Workbook()
b=1

# grab the active worksheet
ws = wb.active
i=1
while i <2:
if ws['A' + str(b)].value == a and ws['A' + str(b)].value != 'x':
return True
break

if ws['A' + str(b)].value == None  :
return False
break


# print('dcm')
b+=1




def count_word():
wb = openpyxl.load_workbook('vocab.xlsx')
# wb = Workbook()



# grab the active worksheet
ws = wb.active
i=1
b=1
c = 0
while i <2:
if ws['A' + str(b)].value != 'x'and ws['A' + str(b)].value != None:
c+=1
if ws['A' + str(b)].value == None  :
break

# print('dcm')
b+=1

return c


def main_vocab():
wb = openpyxl.load_workbook('vocab.xlsx')
# wb = Workbook()



# grab the active worksheet
ws = wb.active




# Data can be assigned directly to cells
# ws['A1'] = 42

# Rows can also be appended
values = []
for var in vars:
if  var.get() != "" :
values.append(var.get())


# a = hihi.get()


# c = count_word()







# word = cell_1.get()
# number = int(input("How many meanings does this word have :"))
# for i in range(1,number+1):
# if i > 1:
# word = "x"
if (check_word(word.get())):
text_variable1.set("Oopss this word you have already learnt")
elif len(values) != 3:
text_variable1.set('You must fill in all 3 boxes!!')
else:
ws.append(values)
global a
if word.get() !='x':
a+=1
hihi.set("Today words: " + str(a))
text_variable1.set('')

# ws.append([word])

# # ws['A1'] = word
# # ws['B1'] = definition
# # ws['C1'] = example
# # sheet.cell(row=2, column=2).value = 2
# # Python types will automatically be convaerted
# # import datetime
# # ws['A2'] = datetime.datetime.now()

# # Save the file

wb.save("vocab.xlsx")
# for i in range(a+1):
for var in vars:
var.set("")
sum1.set('All vocabs: ' + str(count_word()))
if count_word() == 1000:
text_variable.set("Congrats, Your vocabs is now 1000 words!!!")
elif count_word() == 2000:
text_variable.set("dcm 2000 tu kinh z ba:))")
elif count_word() == 3000:
text_variable.set("dcm 3000 tu r nhe dcm ielts 5.0")
elif count_word() == 4000:
text_variable.set("dcm 4000 tu r nhe hehe ielts cx 6.5 roi")
else:
text_variable.set("")
if a == 10:
text_variable.set("Today you have learnt 10 words!!")
elif a == 20:
text_variable.set("Great job!!20 words a day")
elif a == 30:
text_variable.set("dcm 30 tu ngay oke nhe em")
elif a == 50:
text_variable.set("dcm 50 tu ngay oke nhe em")
else:
text_variable.set("")





global a
a = 0
excel = []
word = tk.StringVar()
meaning = tk.StringVar()
example = tk.StringVar()
text_variable = tk.StringVar()
text_variable1 = tk.StringVar()
text_variable2= tk.StringVar()
hihi = tk.IntVar()
sum1 = tk.IntVar()

vars = [word,meaning,example]

columns = ['Word', 'Meaning','Example']
for i in range(3):
label = tk.Label(root,text = columns[i],font = ("Raleway",20),fg = '#826cb2')
label.grid(row=0,column=i)
cell = tk.Entry(root,textvariable = vars[i])
cell.grid(row=1, column=i)

label = tk.Label(root,textvariable = hihi, font = ("Raleway",20))
label.grid(row=3,column=1)
hihi.set("Today words: " + str(a))

label = tk.Label(root,textvariable = sum1, font = ("Raleway",20))
label.grid(row=5,column=1)
sum1.set('All vocabs: ' + str(count_word()))


# cell = tk.Entry(root,textvariable = hihi )
# cell.grid(row=4, column=1)
text= tk.StringVar()
button = tk.Button(root,textvariable= text,command=lambda:main_vocab(), font="Raleway",bg="#20bebe",fg="white",height=2,width= 15)

text.set("New vocab")

text_box = tk.Label (root,textvariable = text_variable,height=6, font = ("Raleway",20), fg = '#1b1b1b' )
text_box.grid(row=7,column=1)
text_box = tk.Label (root,textvariable = text_variable1,height=3, font = ("Raleway",10), fg = '#1b1b1b' )
text_box.grid(row=7,column=1)

# text_box = tk.Label (root,textvariable = text_variable2,height=5, font = ("Raleway",15), fg = '#1b1b1b' )
# text_box.grid(row=7,column=1)
text_variable1.set('Store new english word here , if the word has more than 1 meaning please fill in the word blank x simbol for the second meaning and so on')
button.grid(columnspan=3,column=0,row=6)
canvas = tk.Canvas(root,width = 600 , height= 100)
canvas.grid(columnspan=3)

root.mainloop()