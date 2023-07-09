import tkinter as tk
import PyPDF2
from PIL import Image, ImageTk
from openpyxl import Workbook
import openpyxl
from openpyxl.utils import get_column_letter
import numpy as np
from tkinter import *
from tkinter import font as tkfont
import pyttsx3

engine = pyttsx3.init()
engine.setProperty('rate', 80)
voice_id = "HKEY_LOCAL_MACHINE\SOFTWARE\Microsoft\Speech\Voices\Tokens\TTS_MS_EN-US_ZIRA_11.0"
engine.setProperty('voice', voice_id)


class SampleApp(tk.Tk):

    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)

        self.title_font = tkfont.Font(
            family='Helvetica', size=18, weight="bold", slant="italic")

        # the container is where we'll stack a bunch of frames
        # on top of each other, then the one we want visible
        # will be raised above the others
        container = tk.Frame(self)
        container.pack()
        canvas = tk.Canvas(container, width=1000, height=500)
        canvas.grid(columnspan=3, rowspan=11)

        # create recyclable variables

        self.shared_data = {
        'recyclable_meaning': tk.StringVar(),
        'recyclable_word': tk.StringVar(),
        'recyclable_example': tk.StringVar()

        }

        self.frames = {}
        for F in (StartPage, PageOne, PageTwo):
            page_name = F.__name__
            frame = F(parent=container, controller=self)
            self.frames[page_name] = frame

            # put all of the pages in the same location;
            # the one on the top of the stacking order
            # will be the one that is visible.
            frame.grid(row=0, column=0, sticky="nsew")

        # make the StartPage visible

        self.show_frame("StartPage")
    # put the particular page on top ( visible)

    def show_frame(self, page_name):
        '''Show a frame for the given page name'''
        frame = self.frames[page_name]
        frame.tkraise()


class StartPage(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        # controller help access all the features in class SampleApp
        self.controller = controller
        global listMeaning

        self.tieptheo = 1

        # storing words,meanings,examples

        listMeaning = self.getMeaning()

        e = self.changeWord(minhdepzai	= True)
        self.word = tk.StringVar()
        meaning1 = tk.StringVar()
        meaning2 = tk.StringVar()
        meaning3 = tk.StringVar()
        self.hihi = tk.StringVar()

        var1 = tk.IntVar()
        var2 = tk.IntVar()
        var3 = tk.IntVar()

        # DESIGN

        tb = tk.Label(self, textvariable=self.hihi, font=(
            "Raleway", 20), fg='#826cb2', bg="#20bebe")
        self.hihi.set('Choose the meaning of the word')
        tb.grid(column=1, row=0)

        wordGuess = tk.Label(self, textvariable=self.word, font=(
            "Raleway", 20), fg='#826cb2', bg="#20bebe")
        self.word.set(mainWord)
        wordGuess.grid(column=0, row=0)

        a = [var1, var2, var3]
        self.a = a
        meaning = [meaning1, meaning2, meaning3]
        self.meaning = meaning

        meaningBox1 = tk.Checkbutton(self, textvariable=meaning[0], variable=a[0], font=(
            "Raleway", 20), fg='#000066', bg="#B7B7B7", onvalue=1, offvalue=0)
        meaning1.set(listMeaning[listChoice_index[0]][self.tieptheo][0])
        meaningBox1.grid(column=0, row=2, sticky=tk.W)

        meaningBox2 = tk.Checkbutton(self, textvariable=meaning[1], variable=a[1], font=(
            "Raleway", 20), fg='#000066', bg="#B7B7B7", onvalue=1, offvalue=0)
        meaning2.set(listMeaning[listChoice_index[1]][self.tieptheo][0])
        meaningBox2.grid(column=0, row=3, sticky=tk.W)

        meaningBox3 = tk.Checkbutton(self, textvariable=meaning[2], variable=a[2], font=(
            "Raleway", 20), fg='#000066', bg="#B7B7B7", onvalue=1, offvalue=0)
        meaning3.set(listMeaning[listChoice_index[2]][self.tieptheo][0])
        meaningBox3.grid(column=0, row=4, sticky=tk.W)

        submitButton = tk.Button(self, text='Submit', font=(
            "Raleway", 20), fg='#826cb2', bg="#20bebe", command=lambda: self.checkWord())
        submitButton.grid(column=0, row=7)

        button2 = tk.Button(self, text="Go to Page Two",
                            command=lambda: controller.show_frame("PageTwo"))

        button2.grid(column=0, row=10)
    # a definition helping change words when u r correct

    def changeWord(self, minhdepzai):
        global listChoice_index
        global mainWord_index
        global mainWord

        # pick 3 random word_index from all the word we have leant

        if minhdepzai == False:
            listChoice_index = np.random.randint(
        	    low=1, high=self.count_word(), size=2)
            listChoice_index = np.insert(listChoice_index,0,mainWord_index)
            

        else:



	        mainWord_index	 = np.random.randint(
	            low=1, high=self.count_word(), size=1)
	        mainWord_index = mainWord_index[0]

	        listChoice_index = np.random.randint(
	            low=1, high=self.count_word(), size=2)



	        mainWord = listMeaning[mainWord_index][0]






	        listChoice_index = np.insert(listChoice_index,0,mainWord_index)


	        # choose the first indext we have picked before to become the main word_index

	        # main word


	        

        # shuffle the random list to make sure the anss r not always the first choice of 3

        np.random.shuffle(listChoice_index)
    # a definition helping get words , meanings, examples
    def getMeaning(self):
        wb = openpyxl.load_workbook('vocab.xlsx')
        # wb = Workbook()

        # grab the active worksheet

        listWord = []
        ws = wb.active
        i = 1
        b = 1
        while i < 2:
            listMeaning1 = []
            listExample1 = []
            check = 1
            if ws['A' + str(b)].value != 'x' and ws['A' + str(b)].value != None:
                listMeaning1.append(ws['A' + str(b)].value)
                listExample1.append(ws['B' + str(b)].value)
                listExample1.append(ws['C' + str(b)].value)
                listMeaning1.append(listExample1)





                
               	# this will makes sure that all the meanings of the word are included if the word has 2 or more meanings and examples
                while True:
                    listExample=[]
                    if ws['A' + str(b+check)].value == 'x':

                        listExample.append(ws['B' + str(b+check)].value)
                        listExample.append(ws['C' + str(b+check)].value)
                        listMeaning1.append(listExample)


                    else:
                        listWord.append(listMeaning1)



                        break

                    check += 1

            if ws['A' + str(b)].value == None:
                break

                # print('dcm')
            b += 1


       # the syntax : listword = [word,[meaning1,example1],[meaning2,example2],...[meaningn,examplen]]

        return listWord
    # a definition helping count all the words we have learnt so far
    def count_word(self):
        wb = openpyxl.load_workbook('vocab.xlsx')
        # wb = Workbook()

        # grab the active worksheet
        ws = wb.active
        i = 1
        b = 1
        c = 0
        while i < 2:
            if ws['A' + str(b)].value != 'x' and ws['A' + str(b)].value != None:
                c += 1

            if ws['A' + str(b)].value == None:
                break

                # print('dcm')
            b += 1

        return c
    # a definition helping check wheather the ticked box is correct or not and if it is correct then it will helps change the word
    def checkWord(self):

        for i, j in enumerate(self.a):

            if j.get() == 1:
                if self.meaning[i].get() != listMeaning[mainWord_index][self.tieptheo][0]:
                    self.hihi.set('Incorrect')
                    break
                elif self.meaning[i].get() == listMeaning[mainWord_index][self.tieptheo][0]:
                    self.hihi.set('correct')
                    engine.say(mainWord)
                    engine.runAndWait()




                    # define recyclable variables 
                    self.controller.shared_data['recyclable_example'].set( 'Example: ' + listMeaning[mainWord_index][self.tieptheo][1])
                    self.controller.shared_data['recyclable_meaning'].set('Meaning: ' + listMeaning[mainWord_index][self.tieptheo][0])
                    self.controller.shared_data['recyclable_word'].set(listMeaning[mainWord_index][0])
                    button1 = tk.Button(self, text="See a definition	",
                            command=lambda: self.controller.show_frame("PageOne"))
                    button1.grid(column=0, row=9)

                    





                    if listMeaning[mainWord_index][self.tieptheo][1] != listMeaning[mainWord_index][-1][1]:
                    	self.tieptheo +=1
                    	
                    	d = self.changeWord(minhdepzai = False)
                    	self.word.set(mainWord)
                    	

                    	for k, l in enumerate(self.meaning):

                    		try:

                    			l.set(listMeaning[listChoice_index[k]][self.tieptheo][0])
                    		except:
                    			l.set(listMeaning[listChoice_index[k]][1][0])


                    else:
                    	self.tieptheo = 1
                    	
                    	d = self.changeWord(minhdepzai = True)
                    	self.word.set(mainWord)
                    	for k, l in enumerate(self.meaning):
                    		l.set(listMeaning[listChoice_index[k]][self.tieptheo][0])











                    # change the new word




class PageOne(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        wb = openpyxl.load_workbook('vocab.xlsx')
        ws = wb.active



        label = tk.Label(self, textvariable=self.controller.shared_data['recyclable_word'],
                         font=controller.title_font)
        
        label.pack(side="top", fill="x", pady=10)

        label1 = tk.Label(self, textvariable=  self.controller.shared_data['recyclable_meaning'],
                         font=controller.title_font)
        
        

        label2 = tk.Label(self, textvariable=  self.controller.shared_data['recyclable_example'] ,
                         font=controller.title_font)
        
        label2.pack(side="bottom", fill="y", pady=10)
        label1.pack(side="bottom", fill="x", pady=10)

        button = tk.Button(self, text= 'Return',
                           command=lambda: controller.show_frame("StartPage"))
        button.pack()


class PageTwo(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        label = tk.Label(self, text="This is page 2",
                         font=controller.title_font)
        label.pack(side="top", fill="x", pady=10)
        button = tk.Button(self, text="Go to the start page",
                           command=lambda: controller.show_frame("StartPage"))
        button.pack()


app = SampleApp()
app.mainloop()
